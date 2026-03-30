"""数据导出服务 — 生成 Excel 文件"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.workorder import WorkOrder, StepRecord
from src.models.alert import AlertEvent
from src.models.material_check import MaterialCheck
from src.models.completion_check import CompletionCheck
from src.models.override_log import OverrideLog

_HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_STATUS_MAP = {"running": "进行中", "completed": "已完成", "failed": "失败", "paused": "暂停"}
_SEVERITY_MAP = {"CRITICAL": "严重", "WARN": "警告", "INFO": "信息"}
_RESULT_MAP = {"pass": "通过", "fail": "失败", "skip": "跳过", "mismatch": "不匹配", "match": "匹配"}


def _apply_header(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_cell_style(ws, row_count: int, col_count: int):
    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = alt_fill


def _auto_width(ws, col_count: int, max_width: int = 40):
    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("utf-8")))
        adjusted = min(max(max_len * 0.9 + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _fmt_dt(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt)


async def export_workorders(db: AsyncSession, **filters) -> bytes:
    stmt = select(WorkOrder).order_by(WorkOrder.start_time.desc())
    if filters.get("station_id"):
        stmt = stmt.where(WorkOrder.station_id == filters["station_id"])
    if filters.get("status"):
        stmt = stmt.where(WorkOrder.status == filters["status"])
    if filters.get("start_date"):
        stmt = stmt.where(WorkOrder.start_time >= filters["start_date"])
    if filters.get("end_date"):
        stmt = stmt.where(WorkOrder.start_time <= filters["end_date"])

    result = await db.execute(stmt)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "工单列表"
    headers = ["ID", "序列号", "工位ID", "SOP模板ID", "状态", "操作员ID", "开始时间", "结束时间"]
    _apply_header(ws, headers)

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.sn)
        ws.cell(row=idx, column=3, value=r.station_id)
        ws.cell(row=idx, column=4, value=r.sop_template_id)
        ws.cell(row=idx, column=5, value=_STATUS_MAP.get(r.status, r.status))
        ws.cell(row=idx, column=6, value=r.operator_id)
        ws.cell(row=idx, column=7, value=_fmt_dt(r.start_time))
        ws.cell(row=idx, column=8, value=_fmt_dt(r.end_time))

    _apply_cell_style(ws, len(rows), len(headers))
    _auto_width(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_alerts(db: AsyncSession, **filters) -> bytes:
    stmt = select(AlertEvent).order_by(AlertEvent.created_at.desc())
    if filters.get("station_id"):
        stmt = stmt.where(AlertEvent.station_id == filters["station_id"])
    if filters.get("severity"):
        stmt = stmt.where(AlertEvent.severity == filters["severity"])
    if filters.get("start_date"):
        stmt = stmt.where(AlertEvent.created_at >= filters["start_date"])
    if filters.get("end_date"):
        stmt = stmt.where(AlertEvent.created_at <= filters["end_date"])

    result = await db.execute(stmt)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "报警记录"
    headers = ["ID", "工单ID", "工位ID", "工位编码", "步骤", "告警类型", "严重级别", "消息", "已确认", "时间"]
    _apply_header(ws, headers)

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.workorder_id)
        ws.cell(row=idx, column=3, value=r.station_id)
        ws.cell(row=idx, column=4, value=r.station_code)
        ws.cell(row=idx, column=5, value=r.step_index)
        ws.cell(row=idx, column=6, value=r.alert_type)
        ws.cell(row=idx, column=7, value=_SEVERITY_MAP.get(r.severity, r.severity))
        ws.cell(row=idx, column=8, value=r.message)
        ws.cell(row=idx, column=9, value="是" if r.acknowledged == "1" else "否")
        ws.cell(row=idx, column=10, value=_fmt_dt(r.created_at))

    _apply_cell_style(ws, len(rows), len(headers))
    _auto_width(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_material_checks(db: AsyncSession, **filters) -> bytes:
    stmt = select(MaterialCheck).order_by(MaterialCheck.checked_at.desc())
    if filters.get("workorder_id"):
        stmt = stmt.where(MaterialCheck.workorder_id == filters["workorder_id"])
    if filters.get("start_date"):
        stmt = stmt.where(MaterialCheck.checked_at >= filters["start_date"])
    if filters.get("end_date"):
        stmt = stmt.where(MaterialCheck.checked_at <= filters["end_date"])

    result = await db.execute(stmt)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "物料校验"
    headers = ["ID", "工单ID", "BOM物料项", "识别物料", "结果", "置信度", "检查时间"]
    _apply_header(ws, headers)

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.workorder_id)
        ws.cell(row=idx, column=3, value=r.bom_item)
        ws.cell(row=idx, column=4, value=r.detected_material)
        ws.cell(row=idx, column=5, value=_RESULT_MAP.get(r.result, r.result))
        ws.cell(row=idx, column=6, value=f"{r.confidence:.1%}" if r.confidence else "")
        ws.cell(row=idx, column=7, value=_fmt_dt(r.checked_at))

    _apply_cell_style(ws, len(rows), len(headers))
    _auto_width(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_completion_checks(db: AsyncSession, **filters) -> bytes:
    stmt = select(CompletionCheck).order_by(CompletionCheck.checked_at.desc())
    if filters.get("workorder_id"):
        stmt = stmt.where(CompletionCheck.workorder_id == filters["workorder_id"])
    if filters.get("start_date"):
        stmt = stmt.where(CompletionCheck.checked_at >= filters["start_date"])
    if filters.get("end_date"):
        stmt = stmt.where(CompletionCheck.checked_at <= filters["end_date"])

    result = await db.execute(stmt)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "完工检查"
    headers = ["ID", "工单ID", "结果", "相似度", "缺陷说明", "检查时间"]
    _apply_header(ws, headers)

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.workorder_id)
        ws.cell(row=idx, column=3, value=_RESULT_MAP.get(r.result, r.result))
        ws.cell(row=idx, column=4, value=f"{r.similarity_score:.1%}" if r.similarity_score else "")
        ws.cell(row=idx, column=5, value=r.defects)
        ws.cell(row=idx, column=6, value=_fmt_dt(r.checked_at))

    _apply_cell_style(ws, len(rows), len(headers))
    _auto_width(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_override_logs(db: AsyncSession, **filters) -> bytes:
    stmt = select(OverrideLog).order_by(OverrideLog.created_at.desc())
    if filters.get("workorder_id"):
        stmt = stmt.where(OverrideLog.workorder_id == filters["workorder_id"])
    if filters.get("start_date"):
        stmt = stmt.where(OverrideLog.created_at >= filters["start_date"])
    if filters.get("end_date"):
        stmt = stmt.where(OverrideLog.created_at <= filters["end_date"])

    result = await db.execute(stmt)
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "放行记录"
    headers = ["ID", "工单ID", "步骤序号", "操作员工号", "放行原因", "时间"]
    _apply_header(ws, headers)

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.workorder_id)
        ws.cell(row=idx, column=3, value=r.step_index)
        ws.cell(row=idx, column=4, value=r.operator_badge)
        ws.cell(row=idx, column=5, value=r.reason)
        ws.cell(row=idx, column=6, value=_fmt_dt(r.created_at))

    _apply_cell_style(ws, len(rows), len(headers))
    _auto_width(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
